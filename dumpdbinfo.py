import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, colors
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.hyperlink import Hyperlink
from datetime import datetime, date


def dump_db_info_to_csv(db_name: str, table_dataframes: dict, output_dir: str, sep: str=','):
    """
    Saves each DataFrame in the provided dictionary to a CSV file, organizing the files within a directory 
    named after the database.

    The CSV files are named after the corresponding table names. Each DataFrame is saved to a separate CSV file 
    in the specified output directory, which will include a subdirectory named after the database if not already included.

    Parameters:
    -----------
    db_name : str
        The name of the database, which will be used to create a subdirectory within the output directory.

    table_dataframes : dict of pandas.DataFrame
        A dictionary where each key is a table name and each value is a DataFrame containing the table's column data.
    
    output_dir : str
        The directory where the CSV files will be saved. If the directory does not contain a subdirectory with the 
        name of the database, one will be added.

    sep : str, optional
        Field delimiter for the output file. The default is a comma.

    Returns:
    --------
    None
    """
    # Ensure the output directory includes the database name
    if not output_dir.endswith(db_name):
        output_dir = os.path.join(output_dir, db_name)


    # Ensure the output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Iterate over the dictionary to save each DataFrame to a CSV file
    for table_name, dataframe in table_dataframes.items():
        # Create the CSV file path
        file_path = os.path.join(output_dir, f"{table_name}.csv")
        
        # Save the DataFrame to a CSV file
        dataframe.to_csv(file_path, sep=sep, index=False)


def create_hyperlink(ws, at_cell, sheet_name, cell_ref='A1', display_name=None, font_size=11):
    """
    Creates a hyperlink in a specified cell that links to another cell within the same workbook.

    This function adds a hyperlink to a specified cell in an Excel worksheet (`ws`). The hyperlink points 
    to a cell within another sheet (or the same sheet) within the same workbook. The cell containing the 
    hyperlink is formatted with a blue, underlined font to resemble a standard hyperlink.

    Parameters:
    -----------
    ws : openpyxl.worksheet.worksheet.Worksheet
        The worksheet where the hyperlink will be created.
    
    at_cell : str
        The cell reference (e.g., 'B2') where the hyperlink will be placed in the `ws` worksheet.
    
    sheet_name : str
        The name of the sheet to which the hyperlink will point.
    
    cell_ref : str, optional
        The cell reference within the `sheet_name` sheet that the hyperlink will point to. Default is 'A1'.
    
    display_name : str, optional
        The text to be displayed in the cell containing the hyperlink. If not provided, defaults to the `sheet_name`.

    font_size : int, optional
        The font size to be applied to the cell containing the hyperlink. Default is 11.

    Returns:
    --------
    None
    """
    if display_name is None:
        display_name = sheet_name
    to_location = "'{0}'!{1}".format(sheet_name, cell_ref)
    ws[at_cell].hyperlink = Hyperlink(display=display_name, ref=at_cell, location=to_location)
    ws[at_cell].value = display_name
    ws[at_cell].font = Font(u='single', color=colors.BLUE, size=font_size)


def adjust_column_widths(sheet,  max_width=80):
    """
    Adjusts the width of each column in the Excel sheet based on the maximum width of the data and header values.

    Parameters:
    -----------
    sheet : openpyxl.worksheet.worksheet.Worksheet
        The worksheet where column widths need to be adjusted.

    max_width : int, optional (default=80)
        The maximum allowed width for any column. If the calculated width exceeds this value,
        the column width will be set to this maximum value.
    
    Returns:
    --------
    None
    """
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name

        # Calculate the width required by the header (considering formatting)
        header_length = len(str(col[0].value))
        adjusted_header_length = header_length * 1.5  # Factor to account for bold and larger font size

        # Compare the header length with the lengths of the data values
        for cell in col:
            try:
                cell_length = len(str(cell.value))
                if cell_length > max_length:
                    max_length = cell_length
            except:
                pass
        
        # Use the greater of the header length or data length for column width
        max_length = max(max_length, adjusted_header_length)

        # Adjust the column width and apply the max_width limit
        adjusted_width = min(max_length + 2, max_width)
        sheet.column_dimensions[column].width = adjusted_width


def format_header_cell(cell, font_size=11):
    """
    Formats a header cell with the default styling: white bold text and green background.

    Parameters:
    -----------
    cell : openpyxl.cell.cell.Cell
        The cell to format.
    
    font_size : int, optional
        The font size to be applied to the header cell. Default is 11.

    Returns:
    --------
    None
    """
    cell.font = Font(color="FFFFFF", bold=True, size=font_size + 1)
    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")


def dump_db_info_to_excel(db_name: str, table_dataframes: dict, output_dir: str, include_record_count: bool = False, max_records_per_table: int = 50000):
    """
    Exports data in the provided dictionary to an Excel workbook with a separate sheet for each table's data.

    This function generates an Excel workbook where each table's data is stored in a separate sheet. 
    The first sheet, titled "Tables," serves as an index listing all table names, with hyperlinks to 
    their respective sheets for easy navigation.

    Optionally, a second column can be added to the index sheet, showing the number of records in each table.

    Each table sheet will include a hyperlink in the header row to return to the index sheet.

    Parameters:
    -----------
    db_name : str
        The name of the database, which will be used to name the output Excel file.
    
    table_dataframes : dict of pandas.DataFrame
        A dictionary where each key is a table name and each value is a DataFrame containing the table's column data.
    
    output_dir : str
        The directory where the Excel file will be saved. The function will ensure the directory structure 
        includes a folder named after the database.

    include_record_count : bool, optional
        If True, adds a column to the index sheet showing the number of records in each table. Default is False.

    max_records_per_table : int, optional
        The maximum number of records to include per table in the Excel sheet. Default is 50,000.

    Returns:
    --------
    None
    """

    # Ensure the output directory includes the database name
    if not output_dir.endswith(db_name):
        output_dir = os.path.join(output_dir, db_name)

    # Ensure the output directory exists
    os.makedirs(output_dir, exist_ok=True)

    # Create the Excel workbook
    workbook = Workbook()

    # Default Excel font size if not specified
    standard_font_size = 11  

    # Use the default sheet as the index sheet
    index_sheet = workbook.active
    index_sheet.title = "Tables"
    
    # Add headers to the index sheet
    index_sheet.cell(row=1, column=1, value="Table")
    format_header_cell(index_sheet.cell(row=1, column=1), font_size=standard_font_size)
    if include_record_count:
        index_sheet.cell(row=1, column=2, value="Record Count")
        format_header_cell(index_sheet.cell(row=1, column=2), font_size=standard_font_size)

    # Populate the index sheet with table names (and record counts if enabled)
    for i, (table_name, dataframe) in enumerate(table_dataframes.items(), start=2):
        index_sheet.cell(row=i, column=1, value=table_name)
        if include_record_count:
            index_sheet.cell(row=i, column=2, value=len(dataframe))
    
    # Adjust column width for the index sheet
    adjust_column_widths(index_sheet)

    # Apply a filter to all columns
    index_sheet.auto_filter.ref = index_sheet.dimensions

    for table_name, dataframe in table_dataframes.items():
        # Limit the number of records to max_records_per_table
        limited_dataframe = dataframe.head(max_records_per_table)

        # Create a new sheet with the table name
        sheet = workbook.create_sheet(title=table_name)
        
        # Fijar la primera fila (cabecera)
        sheet.freeze_panes = 'A2'  # Esto congela la primera fila
        
        # Add the DataFrame to the sheet
        for r_idx, row in enumerate(dataframe_to_rows(limited_dataframe, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                # Convert datetime64 and date values to Excel date format
                if isinstance(value, pd.Timestamp):
                    cell_value = value.to_pydatetime()
                elif isinstance(value, date):
                    cell_value = value.date()  
                else:
                    # Ensure the value is converted to a string if it's not a basic data type
                    cell_value = str(value) if not isinstance(value, (int, float, type(None))) else value
                cell = sheet.cell(row=r_idx, column=c_idx, value=cell_value)
                # Apply formatting to header row
                if r_idx == 1:  
                    format_header_cell(cell, font_size=standard_font_size)
        
        # Auto-size columns 
        adjust_column_widths(sheet)

        # Apply a filter to all columns
        sheet.auto_filter.ref = sheet.dimensions

        # Add a hyperlink to return to the "Tables" sheet in the last cell of the header row
        last_col_idx = len(dataframe.columns) + 1
        return_cell = sheet.cell(row=1, column=last_col_idx)
        create_hyperlink(sheet, at_cell=return_cell.coordinate, sheet_name="Tables", cell_ref='A1', display_name="Return to Tables", font_size=standard_font_size+1)

        # Adjust the column width to fit the "Return to Tables" message.
        sheet.column_dimensions[return_cell.column_letter].width = len("Return to Tables") + 2


    # Links are created for each table in the list for easy access to its sheet.
    for i in range(2, index_sheet.max_row + 1):
        create_hyperlink(index_sheet, 'A' + str(i), index_sheet['A' + str(i)].value, cell_ref='A1')

    # Save the workbook to the output directory with the name of the database
    excel_file_path = os.path.join(output_dir, f"{db_name}.xlsx")
    workbook.save(excel_file_path)


